"""
INIT.AI Report Generator.
Produces PDF and XLSX reports using WeasyPrint + openpyxl.
"""

import logging
import io
import os
from datetime import datetime, date
from typing import Dict, Any, Optional
from uuid import UUID

logger = logging.getLogger(__name__)

try:
    from weasyprint import HTML, CSS
    WEASYPRINT_AVAILABLE = True
except ImportError:
    WEASYPRINT_AVAILABLE = False
    logger.warning("WeasyPrint not installed — PDF generation disabled")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False


# ─── HTML Template for PDF ────────────────────────────────────
REPORT_HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"/>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{
    font-family: 'Helvetica Neue', Arial, sans-serif;
    font-size: 11px;
    color: #1a1a1a;
    background: #fff;
    padding: 0;
  }}
  .cover {{
    background: #0f0705;
    color: #fdf0e8;
    padding: 60px 50px;
    min-height: 100vh;
    display: flex;
    flex-direction: column;
    justify-content: center;
  }}
  .cover-logo {{
    font-size: 42px;
    font-weight: 900;
    color: #ff6b2b;
    letter-spacing: -1px;
    margin-bottom: 8px;
  }}
  .cover-subtitle {{ font-size: 13px; color: #c49a82; margin-bottom: 60px; }}
  .cover-title {{ font-size: 28px; font-weight: 700; margin-bottom: 12px; }}
  .cover-city  {{ font-size: 16px; color: #ff6b2b; margin-bottom: 8px; }}
  .cover-date  {{ font-size: 12px; color: #7a5040; }}
  .section {{
    padding: 40px 50px;
    page-break-inside: avoid;
  }}
  .section-title {{
    font-size: 18px;
    font-weight: 700;
    color: #ff6b2b;
    border-bottom: 2px solid #ff6b2b;
    padding-bottom: 8px;
    margin-bottom: 20px;
  }}
  .stats-grid {{
    display: grid;
    grid-template-columns: repeat(4, 1fr);
    gap: 12px;
    margin-bottom: 24px;
  }}
  .stat-box {{
    border: 1px solid #e0d4cc;
    border-radius: 8px;
    padding: 14px;
    text-align: center;
    background: #fdf8f5;
  }}
  .stat-val  {{ font-size: 24px; font-weight: 700; color: #ff3a1a; }}
  .stat-label {{ font-size: 10px; color: #7a5040; text-transform: uppercase; margin-top: 4px; }}
  table {{
    width: 100%;
    border-collapse: collapse;
    font-size: 10px;
    margin-bottom: 20px;
  }}
  th {{
    background: #ff6b2b;
    color: #fff;
    padding: 8px 10px;
    text-align: left;
    font-weight: 600;
  }}
  td {{ padding: 7px 10px; border-bottom: 1px solid #f0e8e0; }}
  tr:nth-child(even) td {{ background: #fdf8f5; }}
  .badge-critical {{ color: #ff3a1a; font-weight: 700; }}
  .badge-high     {{ color: #ff8c00; font-weight: 600; }}
  .badge-moderate {{ color: #d4a017; }}
  .badge-low      {{ color: #22a55e; }}
  .footer {{
    position: fixed;
    bottom: 0;
    left: 0;
    right: 0;
    padding: 10px 50px;
    font-size: 9px;
    color: #7a5040;
    border-top: 1px solid #e0d4cc;
    display: flex;
    justify-content: space-between;
  }}
</style>
</head>
<body>

<div class="cover">
  <div class="cover-logo">INIT.AI</div>
  <div class="cover-subtitle">Urban Heat Island Intelligence Platform · Philippines</div>
  <div class="cover-title">{report_title}</div>
  <div class="cover-city">📍 {city_name}</div>
  <div class="cover-date">Generated: {generated_at} · Data: {data_period}</div>
</div>

<div class="section">
  <div class="section-title">Executive Summary</div>
  <div class="stats-grid">
    <div class="stat-box">
      <div class="stat-val">{avg_lst}°C</div>
      <div class="stat-label">Avg Surface Temp</div>
    </div>
    <div class="stat-box">
      <div class="stat-val">{hotspot_count}</div>
      <div class="stat-label">Active Hotspots</div>
    </div>
    <div class="stat-box">
      <div class="stat-val">{canopy_pct}%</div>
      <div class="stat-label">Tree Canopy Cover</div>
    </div>
    <div class="stat-box">
      <div class="stat-val">{uhi_intensity}°C</div>
      <div class="stat-label">UHI Intensity</div>
    </div>
  </div>
  <p style="color:#4a4a4a;line-height:1.7;font-size:11px">{executive_summary}</p>
</div>

<div class="section">
  <div class="section-title">Active Hotspot Registry</div>
  <table>
    <thead>
      <tr>
        <th>Zone ID</th>
        <th>Location</th>
        <th>LST (°C)</th>
        <th>Severity</th>
        <th>NDVI</th>
        <th>Primary Cause</th>
        <th>Satellite</th>
      </tr>
    </thead>
    <tbody>
      {hotspot_rows}
    </tbody>
  </table>
</div>

<div class="section">
  <div class="section-title">AI Mitigation Recommendations</div>
  {mitigation_content}
</div>

<div class="footer">
  <span>INIT.AI Urban Heat Island Platform · {city_name}</span>
  <span>CONFIDENTIAL — For LGU Internal Use Only</span>
  <span>Page <span class="pagenumber"></span></span>
</div>
</body>
</html>
"""


class ReportGenerator:
    """Generates PDF and XLSX reports from thermal analysis data."""

    def generate_pdf(self, report_data: Dict[str, Any]) -> Optional[bytes]:
        """Generate a PDF report and return bytes."""
        if not WEASYPRINT_AVAILABLE:
            logger.warning("WeasyPrint not available — returning mock PDF bytes")
            return b"%PDF-1.4 mock report content"

        hotspot_rows = ""
        for h in report_data.get("hotspots", []):
            sev_class = f"badge-{h.get('severity','moderate')}"
            hotspot_rows += f"""
            <tr>
              <td><b>{h.get('zone_id','')}</b></td>
              <td>{h.get('barangay_name','')}</td>
              <td style="font-weight:700;color:#ff3a1a">{h.get('lst','')}</td>
              <td class="{sev_class}">{str(h.get('severity','')).upper()}</td>
              <td>{h.get('ndvi','')}</td>
              <td>{h.get('cause','')}</td>
              <td>{h.get('satellite','')}</td>
            </tr>"""

        mitigation_items = ""
        for m in report_data.get("mitigations", []):
            mitigation_items += f"""
            <div style="margin-bottom:14px;padding:12px;border:1px solid #e0d4cc;border-radius:6px">
              <div style="font-weight:700;color:#ff6b2b;margin-bottom:4px">{m.get('title','')}</div>
              <div style="color:#4a4a4a">{m.get('description','')}</div>
              <div style="color:#22a55e;margin-top:4px;font-weight:600">
                Est. Impact: {m.get('impact','')} | Cost: {m.get('cost','')} | Timeline: {m.get('timeline','')}
              </div>
            </div>"""

        html_content = REPORT_HTML_TEMPLATE.format(
            report_title   = report_data.get("title", "UHI Assessment Report"),
            city_name      = report_data.get("city_name", "Quezon City"),
            generated_at   = datetime.now().strftime("%B %d, %Y %H:%M PHT"),
            data_period    = report_data.get("data_period", "May 2025"),
            avg_lst        = report_data.get("avg_lst", "38.4"),
            hotspot_count  = report_data.get("hotspot_count", 12),
            canopy_pct     = report_data.get("canopy_pct", "18.3"),
            uhi_intensity  = report_data.get("uhi_intensity", "4.8"),
            executive_summary = report_data.get("executive_summary", ""),
            hotspot_rows   = hotspot_rows,
            mitigation_content = mitigation_items,
        )

        try:
            pdf_bytes = HTML(string=html_content).write_pdf()
            return pdf_bytes
        except Exception as e:
            logger.error(f"PDF generation failed: {e}")
            return None

    def generate_xlsx(self, report_data: Dict[str, Any]) -> Optional[bytes]:
        """Generate an Excel report and return bytes."""
        if not OPENPYXL_AVAILABLE:
            logger.warning("openpyxl not available")
            return None

        wb = openpyxl.Workbook()

        # ── Summary Sheet ──────────────────────────────────
        ws = wb.active
        ws.title = "Summary"
        header_fill = PatternFill("solid", fgColor="FF6B2B")
        header_font = Font(bold=True, color="FFFFFF", size=12)

        ws.merge_cells("A1:F1")
        ws["A1"] = f"INIT.AI UHI Report — {report_data.get('city_name','Quezon City')}"
        ws["A1"].font = Font(bold=True, size=16, color="FF3A1A")

        ws["A3"] = "Metric"
        ws["B3"] = "Value"
        for cell in [ws["A3"], ws["B3"]]:
            cell.fill = header_fill
            cell.font = header_font

        stats = [
            ("Average Surface Temp", f"{report_data.get('avg_lst', 38.4)}°C"),
            ("Active Hotspots",      report_data.get("hotspot_count", 12)),
            ("Tree Canopy Cover",    f"{report_data.get('canopy_pct', 18.3)}%"),
            ("UHI Intensity",        f"{report_data.get('uhi_intensity', 4.8)}°C"),
            ("Report Generated",     datetime.now().strftime("%Y-%m-%d %H:%M")),
        ]
        for i, (label, value) in enumerate(stats, start=4):
            ws[f"A{i}"] = label
            ws[f"B{i}"] = value

        # ── Hotspots Sheet ─────────────────────────────────
        ws2 = wb.create_sheet("Hotspots")
        cols = ["Zone ID","Location","LST (°C)","Severity","NDVI","Impervious %","Cause","Satellite"]
        for j, col in enumerate(cols, 1):
            cell = ws2.cell(row=1, column=j, value=col)
            cell.fill = header_fill
            cell.font = header_font

        for i, h in enumerate(report_data.get("hotspots", []), start=2):
            ws2.cell(row=i, column=1, value=h.get("zone_id"))
            ws2.cell(row=i, column=2, value=h.get("barangay_name"))
            ws2.cell(row=i, column=3, value=float(h.get("lst", 0)))
            ws2.cell(row=i, column=4, value=h.get("severity","").upper())
            ws2.cell(row=i, column=5, value=float(h.get("ndvi", 0)))
            ws2.cell(row=i, column=6, value=float(h.get("impervious_pct", 0)))
            ws2.cell(row=i, column=7, value=h.get("cause"))
            ws2.cell(row=i, column=8, value=h.get("satellite"))

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()


# Singleton
report_generator = ReportGenerator()
